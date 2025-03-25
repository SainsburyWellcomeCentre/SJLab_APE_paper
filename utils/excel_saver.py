from openpyxl import Workbook
import matplotlib.pyplot as plt
from matplotlib.collections import PathCollection


def save_figure_data_to_excel(fig, filename='figure_data.xlsx'):
    # Create a new workbook and remove the default sheet
    wb = Workbook()
    wb.remove(wb.active)

    # Extract data from all axes in the figure
    for i, ax in enumerate(fig.get_axes(), start=1):
        # Create a new sheet for each axis
        sheet = wb.create_sheet(title=f'Axis {i}')
        sheet.append(['Line Label', 'X Data', 'Y Data'])

        for line in ax.get_lines():
            x_data = line.get_xdata()
            y_data = line.get_ydata()
            label = line.get_label()
            for x, y in zip(x_data, y_data):
                sheet.append([label, x, y])

        # Handle scatter plots
        for collection in ax.collections:
            if isinstance(collection, PathCollection):
                offsets = collection.get_offsets()
                label = collection.get_label()
                for offset in offsets:
                    x, y = offset
                    sheet.append([label, x, y])


    # Save the workbook
    wb.save(filename)
    print(f"Data has been saved to {filename}")

    return


def scatter_to_excel(fig, filename='scatter_data.xlsx'):
    # Create a new workbook and remove the default sheet
    wb = Workbook()
    wb.remove(wb.active)

    # Extract data from all axes in the figure
    for i, ax in enumerate(fig.get_axes(), start=1):
        # Create a new sheet for each axis
        sheet = wb.create_sheet(title=f'Axis {i}')
        sheet.append(['Line Label', 'X Data', 'Y Data'])

        for collection in ax.collections:
            if isinstance(collection, PathCollection):
                offsets = collection.get_offsets()
                label = collection.get_label()
                for offset in offsets:
                    x, y = offset
                    sheet.append([label, x, y])

    # Save the workbook
    wb.save(filename)
    print(f"Data has been saved to {filename}")

    return